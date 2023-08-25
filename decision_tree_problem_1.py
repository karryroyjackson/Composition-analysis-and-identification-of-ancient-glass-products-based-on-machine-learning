import openpyxl
import numpy as np

sample_count = 56


def generate_tree_partitions():  # 按纹饰、类型建树
    wb = openpyxl.load_workbook('附件.xlsx')
    ws = wb['表单1']
    tree_partitions = [[] for _ in range(6)]
    # tree_partitions = [[] for _ in range(16)]
    # tree_partitions = [[] for _ in range(24)]
    for row in range(2, 58):
        data1 = ws.cell(row=row, column=1).value
        data2 = ws.cell(row=row, column=2).value
        data3 = ws.cell(row=row, column=3).value
        data4 = ws.cell(row=row, column=4).value
        data5 = ws.cell(row=row, column=5).value
        tup = (data1, data2, data3, data4, data5)
        # manual decision trees
        # # 按纹饰、类型建树
        if data2 == 'A':
            if data3 == '高钾':
                tree_partitions[0].append(tup)
            else:  # '铅钡'
                tree_partitions[1].append(tup)
        elif data2 == 'B':
            if data3 == '高钾':
                tree_partitions[2].append(tup)
            else:  # '铅钡'
                tree_partitions[3].append(tup)
        else:  # 'C'
            if data3 == '高钾':
                tree_partitions[4].append(tup)
            else:  # '铅钡'
                tree_partitions[5].append(tup)

        # # 按类型、颜色建树
        # match data3:
        #     case "高钾":
        #         match data4:
        #             case "黑":
        #                 tree_partitions[0].append(tup)
        #             case "蓝绿":
        #                 tree_partitions[1].append(tup)
        #             case "绿":
        #                 tree_partitions[2].append(tup)
        #             case "浅蓝":
        #                 tree_partitions[3].append(tup)
        #             case "浅绿":
        #                 tree_partitions[4].append(tup)
        #             case "深蓝":
        #                 tree_partitions[5].append(tup)
        #             case "深绿":
        #                 tree_partitions[6].append(tup)
        #             case "紫":
        #                 tree_partitions[7].append(tup)
        #     case "铅钡":
        #         match data4:
        #             case "黑":
        #                 tree_partitions[8].append(tup)
        #             case "蓝绿":
        #                 tree_partitions[9].append(tup)
        #             case "绿":
        #                 tree_partitions[10].append(tup)
        #             case "浅蓝":
        #                 tree_partitions[11].append(tup)
        #             case "浅绿":
        #                 tree_partitions[12].append(tup)
        #             case "深蓝":
        #                 tree_partitions[13].append(tup)
        #             case "深绿":
        #                 tree_partitions[14].append(tup)
        #             case "紫":
        #                 tree_partitions[15].append(tup)

        # # 按纹饰、颜色建树
        # match data2:
        #     case "A":
        #         match data4:
        #             case "黑":
        #                 tree_partitions[0].append(tup)
        #             case "蓝绿":
        #                 tree_partitions[1].append(tup)
        #             case "绿":
        #                 tree_partitions[2].append(tup)
        #             case "浅蓝":
        #                 tree_partitions[3].append(tup)
        #             case "浅绿":
        #                 tree_partitions[4].append(tup)
        #             case "深蓝":
        #                 tree_partitions[5].append(tup)
        #             case "深绿":
        #                 tree_partitions[6].append(tup)
        #             case "紫":
        #                 tree_partitions[7].append(tup)
        #     case "B":
        #         match data4:
        #             case "黑":
        #                 tree_partitions[8].append(tup)
        #             case "蓝绿":
        #                 tree_partitions[9].append(tup)
        #             case "绿":
        #                 tree_partitions[10].append(tup)
        #             case "浅蓝":
        #                 tree_partitions[11].append(tup)
        #             case "浅绿":
        #                 tree_partitions[12].append(tup)
        #             case "深蓝":
        #                 tree_partitions[13].append(tup)
        #             case "深绿":
        #                 tree_partitions[14].append(tup)
        #             case "紫":
        #                 tree_partitions[15].append(tup)
        #     case "C":
        #         match data4:
        #             case "黑":
        #                 tree_partitions[16].append(tup)
        #             case "蓝绿":
        #                 tree_partitions[17].append(tup)
        #             case "绿":
        #                 tree_partitions[18].append(tup)
        #             case "浅蓝":
        #                 tree_partitions[19].append(tup)
        #             case "浅绿":
        #                 tree_partitions[20].append(tup)
        #             case "深蓝":
        #                 tree_partitions[21].append(tup)
        #             case "深绿":
        #                 tree_partitions[22].append(tup)
        #             case "紫":
        #                 tree_partitions[23].append(tup)

    return tree_partitions


def statistics_of_tree_partitions(tree_partitions):
    stat_info = []
    for i in range(6):
        # for i in range(16):
        # for i in range(24):
        lst = tree_partitions[i]
        if not lst:  # if empty
            stat_info.append(("*", "*", "*"))
            continue
        fenghua_count = 0
        for item in lst:
            if item[4] == '风化':
                fenghua_count += 1
        weifenghua_count = len(lst) - fenghua_count
        stat_info.append((fenghua_count / len(lst), weifenghua_count / len(lst), len(lst) / sample_count))
    return stat_info


def gini(p):
    return p * (1 - p) + (1 - p) * (1 - (1 - p))


def entropy(p):
    if p == 0.0 or p == 1.0:
        return 0.0
    return -p * np.log2(p) - (1 - p) * np.log2((1 - p))


def error(p):
    return 1 - np.max([p, 1 - p])


def weighted_sum(stat_info, cond):
    s = 0
    for item in stat_info:
        if item == ("*", "*", "*"):
            continue

        p = item[0]
        match cond:
            case "gini":
                self_impurity = gini(p)
            case "entropy":
                self_impurity = entropy(p)
                # print(p, self_impurity)
            case _:  # "error"
                self_impurity = error(p)
        s += item[2] * self_impurity
    return s


def main():
    lst = generate_tree_partitions()
    for item in lst:
        print(item)
    stat_info = statistics_of_tree_partitions(lst)
    print(stat_info)
    # print(weighted_sum(stat_info, cond='gini'))
    gini_index, entropy_index, error_index = weighted_sum(stat_info, cond='gini'), weighted_sum(stat_info,
                                                                                                cond='entropy'), weighted_sum(
        stat_info, cond='error')
    print("gini_index =", gini_index)
    print("entropy_index =", entropy_index)
    print("error_index =", error_index)


if __name__ == "__main__":
    main()
