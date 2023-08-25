import openpyxl
import numpy as np
import random
from sklearn.metrics import classification_report
from sklearn.cluster import KMeans
from sklearn.metrics import silhouette_samples

import math

from matplotlib import pyplot as plt

from sklearn.neighbors import KNeighborsClassifier
from sklearn.naive_bayes import GaussianNB
from sklearn.linear_model import LogisticRegression
from sklearn.svm import SVC
from sklearn.tree import DecisionTreeClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.neural_network import MLPClassifier

# 0 for 高钾，1 for 铅钡


def glass_index_to_type(file_name):
    wb = openpyxl.load_workbook(file_name)
    ws = wb["表单1"]
    index_to_type = dict()
    for row in range(2, ws.max_row + 1):
        glass_id = ws.cell(row=row, column=1).value
        glass_type = ws.cell(row=row, column=3).value
        index_to_type[glass_id] = glass_type
    wb.close()
    return index_to_type


def glass_index_to_weathered(file_name):
    wb = openpyxl.load_workbook(file_name)
    ws = wb["表单1"]
    index_to_weathered = dict()
    for row in range(2, ws.max_row + 1):
        glass_id = ws.cell(row=row, column=1).value
        glass_weather_extent = ws.cell(row=row, column=5).value
        index_to_weathered[glass_id] = glass_weather_extent
    wb.close()
    return index_to_weathered


def shuffled_labeled_data_from_file(file_name):
    index_to_type = glass_index_to_type(file_name)
    print(index_to_type)
    wb = openpyxl.load_workbook(file_name)
    ws = wb['表单2']
    data_lst = [[] for _ in range(2, 71)]
    # label_lst = []
    for row in range(2, 71):
        if "严重风化" in ws.cell(row=row, column=1).value:
            # print("delete one row due to pollution")
            continue
        s = 0.0
        for col in range(2, 16):
            c = ws.cell(row=row, column=col)
            data = c.value
            if data is None:
                c.value = 0.0
            s = s + c.value
        if not 85 <= s <= 105:
            continue
        # append possibly several features
        for col in range(1, 16):
            # append a single feature
            data = ws.cell(row=row, column=col).value
            if col == 1:
                data = data[:2]
            data_lst[row - 2].append(data)

        # append labels
        name = ws.cell(row=row, column=1).value[:2]
        label = index_to_type[name]
        data_lst[row - 2].append(label)
        # print(name, label)

    for item in data_lst:
        if not item:
            data_lst.remove(item)

    high_k_data_lst = [item for item in data_lst if item[-1] == '高钾']
    Pb_Ba_data_lst = [item for item in data_lst if item[-1] == '铅钡']

    random.shuffle(data_lst)
    wb.close()
    # data_lst is shuffled, but the latter two are not
    return data_lst, high_k_data_lst, Pb_Ba_data_lst


def lists_via_proportional_split(lst, ratio):
    counter = int(len(lst) * ratio)
    return lst[:counter], lst[counter:]


def labeled_data_from_file_for_prediction(file_name):
    wb = openpyxl.load_workbook(file_name)
    ws = wb['表单3']
    feature_lst = [[] for _ in range(ws.max_row - 1)]
    for row in range(2, ws.max_row + 1):
        for col in range(3, 17):
            data = ws.cell(row=row, column=col).value
            if data is None:
                data = 0.0
            feature_lst[row - 2].append(data)

    wb.close()
    return feature_lst


def kmeans_clustering_analysis(data_lst, n_clusters):
    km = KMeans(n_clusters=n_clusters, init='k-means++', n_init=10, max_iter=300, tol=1e-04, random_state=0)
    y_km = km.fit_predict(data_lst)
    # print("*************")
    # print("data_lst")
    # print(np.array(data_lst))
    # print(len(data_lst))
    # print("y_km")
    # print(y_km)
    # print(len(y_km))
    # print("*************")
    silhouette_vals = silhouette_samples(data_lst, y_km, metric='euclidean')
    silhouette_avg = np.mean(silhouette_vals)
    return km.inertia_, silhouette_avg, km.cluster_centers_


def draw_elbow_graph(X):
    distortions = []
    for i in range(1, 11):
        km = KMeans(n_clusters=i, init='k-means++', n_init=10, max_iter=300, random_state=0)
        km.fit(X)
        distortions.append(km.inertia_)
    plt.plot(range(1, 11), distortions, marker='o')
    plt.xlabel('Number of clusters')
    plt.ylabel('Distortion')
    plt.tight_layout()
    plt.show()


def draw_plots(data_lst):
    i = 0
    label_lst = ["SiO$_2$", r"Na$_2$O", r"K$_2$O", "CaO", "MgO", r"Al$_2$O$_3$", r"Fe$_2$O$_3$",
                 "CuO", "PbO", "BaO", r"P$_2$O$_5$", "SrO", r"SnO$_2$", r"SO$_2$"]
    for j in range(1, len(data_lst[0])):
        x_lst = [item[i] for item in data_lst]
        y_lst = [item[j] for item in data_lst]
        print(x_lst)
        print(y_lst)
        plt.plot(x_lst, y_lst, marker='o')
        plt.xlabel(label_lst[0])
        plt.ylabel(label_lst[j])
        plt.tight_layout()
        plt.title("Relevance between " + label_lst[j] + " and " + label_lst[i])
        plt.show()


def draw_single_plot(data_lst, i, j):
    label_lst = ["SiO$_2$", r"Na$_2$O", r"K$_2$O", "CaO", "MgO", r"Al$_2$O$_3$", r"Fe$_2$O$_3$",
                 "CuO", "PbO", "BaO", r"P$_2$O$_5$", "SrO", r"SnO$_2$", r"SO$_2$"]
    x_lst = [item[i] for item in data_lst]
    y_lst = [item[j] for item in data_lst]
    print(x_lst)
    print(y_lst)
    plt.plot(x_lst, y_lst, marker='o')
    plt.xlabel(label_lst[0])
    plt.ylabel(label_lst[j])
    plt.tight_layout()
    plt.title("Relevance between " + label_lst[j] + " and " + label_lst[i])
    plt.show()


def draw_subplots_against_1st_item_together(data_lst):
    i = 0
    label_lst = ["SiO$_2$", r"Na$_2$O", r"K$_2$O", "CaO", "MgO", r"Al$_2$O$_3$", r"Fe$_2$O$_3$",
                 "CuO", "PbO", "BaO", r"P$_2$O$_5$", "SrO", r"SnO$_2$", r"SO$_2$"]
    for j in range(1, len(data_lst[0])):
        plt.subplot(4, 4, j)
        x_lst = [item[i] for item in data_lst]
        y_lst = [item[j] for item in data_lst]
        print(x_lst)
        print(y_lst)
        plt.plot(x_lst, y_lst, marker='o')
        plt.xlabel(label_lst[0])
        plt.ylabel(label_lst[j])
        plt.tight_layout()
    # plt.title("Relevance between other elements and Silicon")
    plt.show()


def draw_subplots_4_pairwise_analysis_together(data_lst):
    label_lst = ["SiO$_2$", r"Na$_2$O", r"K$_2$O", "CaO", "MgO", r"Al$_2$O$_3$", r"Fe$_2$O$_3$",
                 "CuO", "PbO", "BaO", r"P$_2$O$_5$", "SrO", r"SnO$_2$", r"SO$_2$"]
    n = 0
    for i in range(0, len(data_lst[0]) - 1):
        for j in range(i + 1, len(data_lst[0])):
            plt.subplot(4, 4, n + 1)
            n += 1
            x_lst = [item[i] for item in data_lst]
            y_lst = [item[j] for item in data_lst]
            print(x_lst)
            print(y_lst)
            plt.scatter(x_lst, y_lst, marker='o')
            plt.xlabel(label_lst[i])
            plt.ylabel(label_lst[j])
            plt.tight_layout()
            # plt.title("Relevance between other elements and Silicon")
            if n % 12 == 0:
                n = 0
                plt.show()


def high_dim_distance(p1: tuple, p2: tuple, digits=2):
    a = map(lambda x: (x[0] - x[1])**2, zip(p1, p2))
    return round(math.sqrt(sum(a)), digits)


def nearest_centers_to_new_data_points(data_lst, center_lst):
    location_lst = []
    for data in data_lst:
        distance_lst = [high_dim_distance(data, center_lst[i]) for i in range(len(center_lst))]
        print("distance_lst: ", distance_lst)
        location = distance_lst.index(min(distance_lst))
        location_lst.append(location)
    return location_lst


def average_of_data_lists(data_lst):
    data_lists_avg = [round(sum([item[i] for item in data_lst]) / len(data_lst), 2)
                      for i in range(len(data_lst[0]))]
    data_lists_std = [round(sum([(item[i] - data_lists_avg[i])**2 for item in data_lst]) / (len(data_lst) - 1), 2)
                      for i in range(len(data_lst[0]))]
    return data_lists_avg, data_lists_std


def all_acc_in_k_fold_validation(model, k, train_data, train_targets):
    # k = 4
    num_val_samples = len(train_data) // k
    num_epochs = 100
    all_accuracies = []
    for i in range(k):
        print(f"Processing fold #{i}")
        val_data = train_data[i * num_val_samples: (i + 1) * num_val_samples]
        val_targets = train_targets[i * num_val_samples: (i + 1) * num_val_samples]
        partial_train_data = np.concatenate([train_data[:i * num_val_samples],
                                             train_data[(i + 1) * num_val_samples:]],
                                            axis=0)
        partial_train_targets = np.concatenate([train_targets[:i * num_val_samples],
                                                train_targets[(i + 1) * num_val_samples:]],
                                               axis=0)
        # model.fit(partial_train_data, partial_train_targets, epochs=num_epochs, batch_size=16, verbose=0)
        model.fit(partial_train_data, partial_train_targets)
        # acc = model.score(val_data, val_targets, verbose=0)
        acc = model.score(val_data, val_targets)
        all_accuracies.append(acc)
    return all_accuracies


def main():
    file_name = '附件.xlsx'
    data_lst, high_k_data_lst, Pb_Ba_data_lst = shuffled_labeled_data_from_file(file_name)
    print("data_lst")
    print(data_lst)
    print("high_k_data_lst")
    print(high_k_data_lst)
    index_to_weathered = glass_index_to_weathered(file_name)
    high_k_weathered_data_lst = [item[1:-1] for item in high_k_data_lst if index_to_weathered[item[0]] == '风化']
    high_k_non_weathered_data_lst = [item[1:-1] for item in high_k_data_lst if index_to_weathered[item[0]] != '风化']

    high_k_weathered_avg, high_k_weathered_std = average_of_data_lists(high_k_weathered_data_lst)
    high_k_non_weathered_avg, high_k_non_weathered_std = average_of_data_lists(high_k_non_weathered_data_lst)

    # print("high_k_weathered_data_lst")
    # for item in high_k_weathered_data_lst:
    #     print(item)
    #
    # print("high_k_non_weathered_data_lst")
    # for item in high_k_non_weathered_data_lst:
    #     print(item)

    high_k_weathered_data_tuple_lst = [tuple(item) for item in high_k_weathered_data_lst]
    high_k_non_weathered_data_tuple_lst = [tuple(item) for item in high_k_non_weathered_data_lst]

    # 高钾风化样本中的化学成分两两分析
    # draw_subplots_4_pairwise_analysis_together(high_k_weathered_data_tuple_lst)

    # 高钾无风化样本中的化学成分两两分析
    draw_subplots_4_pairwise_analysis_together(high_k_non_weathered_data_tuple_lst)

    high_k_weathered_data_tuple_lst.sort(key=lambda x: x[0], reverse=False)  # sort according to Si
    print("high_k_weathered_data_tuple_lst (sorted)")
    for item in high_k_weathered_data_tuple_lst:
        print(item)
    print()

    # draw_plots(high_k_weathered_data_tuple_lst)
    # draw_subplots_against_1st_item_together(high_k_weathered_data_tuple_lst)

    print("high_k_weathered_avg")
    print(high_k_weathered_avg)
    print("high_k_non_weathered_avg")
    print(high_k_non_weathered_avg)
    print("high_k_weathered_std")
    print(high_k_weathered_std)
    print("high_k_non_weathered_std")
    print(high_k_non_weathered_std)

    print("Pb_ba_data_lst")
    print(Pb_Ba_data_lst)
    Pb_Ba_weathered_data_lst = [item[1:-1] for item in Pb_Ba_data_lst if index_to_weathered[item[0]] == '风化']
    Pb_Ba_non_weathered_data_lst = [item[1:-1] for item in Pb_Ba_data_lst if index_to_weathered[item[0]] != '风化']


    Pb_Ba_weathered_avg, Pb_Ba_weathered_std = average_of_data_lists(Pb_Ba_weathered_data_lst)
    Pb_Ba_non_weathered_avg, Pb_Ba_non_weathered_std = average_of_data_lists(Pb_Ba_non_weathered_data_lst)
    print("Pb_Ba_weathered_avg")
    print(Pb_Ba_weathered_avg)
    print("Pb_Ba_non_weathered_avg")
    print(Pb_Ba_non_weathered_avg)
    print("Pb_Ba_weathered_std")
    print(Pb_Ba_weathered_std)
    print("Pb_Ba_non_weathered_std")
    print(Pb_Ba_non_weathered_std)

    Pb_Ba_weathered_data_tuple_lst = [tuple(item) for item in Pb_Ba_weathered_data_lst]
    Pb_Ba_non_weathered_data_tuple_lst = [tuple(item) for item in Pb_Ba_non_weathered_data_lst]

    # 铅钡风化样本中的化学成分两两分析
    # draw_subplots_4_pairwise_analysis_together(Pb_Ba_weathered_data_tuple_lst)

    # 铅钡无风化样本中的化学成分两两分析
    # draw_subplots_4_pairwise_analysis_together(Pb_Ba_non_weathered_data_tuple_lst)

    Pb_Ba_weathered_data_tuple_lst.sort(key=lambda x: x[0], reverse=True)
    print("Pb_Ba_weathered_data_tuple_lst (sorted)")
    for item in Pb_Ba_weathered_data_tuple_lst:
        print(item)
    print()
    # draw_subplots_against_1st_item_together(Pb_Ba_weathered_data_tuple_lst)
    print("****************")
    feature_lst = [item[1:-1] for item in data_lst]
    label_lst = [item[-1] for item in data_lst]
    label_lst = [0 if item == '高钾' else 1 for item in label_lst]
    high_k_data_lst = [item[1:-1] for item in high_k_data_lst]
    Pb_Ba_data_lst = [item[1:-1] for item in Pb_Ba_data_lst]

    train_features, test_features = lists_via_proportional_split(feature_lst, 0.8)
    train_labels, test_labels = lists_via_proportional_split(label_lst, 0.8)
    print("train_features")
    print(train_features)
    print("test_features")
    print(test_features)

    trainX = np.array(train_features)
    trainY = np.array(train_labels)
    testX = np.array(test_features)
    testY = np.array(test_labels)

    name_to_model = {
        "knn": KNeighborsClassifier(n_neighbors=1),
        "naive_bayes": GaussianNB(),
        "logit": LogisticRegression(solver="lbfgs", multi_class="auto"),
        "svm": SVC(kernel="rbf", gamma="auto", probability=True),
        "decision_tree": DecisionTreeClassifier(),
        "random_forest": RandomForestClassifier(n_estimators=100),
        "mlp": MLPClassifier()
    }

    # model = name_to_model["knn"]
    model = name_to_model["naive_bayes"]
    # model = name_to_model["logit"]
    # model = name_to_model["svm"]
    # model = name_to_model["decision_tree"]
    # model = name_to_model["random_forest"]
    # model = name_to_model["mlp"]

    model.fit(trainX, trainY)
    acc = model.score(testX, testY)
    print("acc in main():", acc, "\n")

    fold_count = 10
    all_accuracies = all_acc_in_k_fold_validation(model=model, k=fold_count, train_data=trainX, train_targets=trainY)
    print("all_accuracies")
    print(all_accuracies)
    avg_acc = round(sum(all_accuracies) / len(all_accuracies), 2)
    print(f"avg_acc in {fold_count} folds:", avg_acc)

    # 预测并输出一份分类结果报告
    print("评估模型效果...")
    predictions = model.predict(testX)
    print(classification_report(testY, predictions))

    feature_lst_4_predictions = labeled_data_from_file_for_prediction(file_name)
    X_to_predict = np.array(feature_lst_4_predictions)
    predicted_prob = model.predict_proba(X_to_predict[:, :])
    print(predicted_prob)
    predicted_classes = model.predict_proba(X_to_predict[:, :]).argmax(axis=1)
    print("predicted_classes", predicted_classes)

    predicted_high_k_data_lst = [X_to_predict[index]
                                 for index in range(len(predicted_classes)) if predicted_classes[index] == 0]
    predicted_Pb_Ba_data_lst = [X_to_predict[index]
                                for index in range(len(predicted_classes)) if predicted_classes[index] == 1]

    print("predicted_high_k_data_lst:", predicted_high_k_data_lst)
    print("predicted_Pb_Ba_data_lst:", predicted_Pb_Ba_data_lst)

    # high K and Pb+Ba
    print(high_k_data_lst)
    print(Pb_Ba_data_lst)

    # as to high k
    # draw_elbow_graph(high_k_data_lst)

    inertia, sil_avg, cluster_centers = kmeans_clustering_analysis(high_k_data_lst, 2)
    print("for high K:")
    print("inertia: ", round(inertia, 2), "sil_average", sil_avg)
    print("cluster centers: ", cluster_centers)
    proper_centers = nearest_centers_to_new_data_points(predicted_high_k_data_lst, cluster_centers)
    print("proper_centers (subclasses)", proper_centers)

    # as to high Pb-Ba
    # draw_elbow_graph(Pb_Ba_data_lst)

    inertia, sil_avg, cluster_centers = kmeans_clustering_analysis(Pb_Ba_data_lst, 2)
    print("for Pb-Ba:")
    print("inertia: ", round(inertia, 2), "sil_average", sil_avg)
    print("cluster centers: ", cluster_centers)
    proper_centers = nearest_centers_to_new_data_points(predicted_Pb_Ba_data_lst, cluster_centers)
    print("proper_centers (subclasses)", proper_centers)


if __name__ == "__main__":
    main()
